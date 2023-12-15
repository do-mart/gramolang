"""
Completion from an excel (and eventually other types of) sheet

TODO: Create template.xlsx with help, and docstring to agent.setters...
TODO: Fix ResourceWarning: unclosed <ssl.SSLSocket for unittest testing
"""

from typing import Sequence
from logging import getLogger

from datetime import datetime
from pathlib import Path
from threading import Lock
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.utils.exceptions import IllegalCharacterError

from .common import (
    COMMENT_CHAR,
    now_delta, write_now_delta, join_none, rmark,
    write_exception, write_error)
from .wraipi import APIWrapper
from .command import BaseEmptyCommand
from .chat import Chat, CompleteCommand


# Separator for chat id
CHAT_ID_SEP = '-'

# Font Style for Error Cell
COMPLETE_FONT = Font(color='0000FF')
ERROR_FONT = Font(bold=True, color='FF0000')

# Logging
module_logger = getLogger(__name__)


def _modify_cell(cell, lock, value, error=False):
    lock.acquire()
    try:
        if error: cell.font = ERROR_FONT
        else: cell.font = COMPLETE_FONT
        cell.value = value
    except Exception as e:
        if error and isinstance(e, IllegalCharacterError):
            # Note: models can output illegal cell character for Excel (especially
            # at high temperature setting). This handler ensures that the error
            # message will be printed in the cell without raising another
            # exception. The original output should raise an exception and won't
            # be printed in the cell, only the formated error message.
            cell.value = ILLEGAL_CHARACTERS_RE.sub(r'', value)
        else:
            raise e
    finally:
        lock.release()


def _complete_col(
        name, command_names, col, api_keys, api_key_files, model,
        timeout, retries, lock, call_id):

    start = datetime.now()
    logger = module_logger.getChild(_complete_col.__name__)
    logger.info(f"Starting chat completion of {name}")

    exceptions = []
    chat = Chat(api_keys=api_keys, api_key_files=api_key_files)
    chat.model(model)
    if timeout is not None: chat.timeout = timeout
    if retries is not None: chat.retries = retries

    # Cell loop
    idx = -1
    for j, cell in enumerate(col[1:]):

        # Skip line of commented command
        if command_names[j].startswith(COMMENT_CHAR): continue

        try:

            # Create command instance
            if issubclass(chat.commands[command_names[j]], BaseEmptyCommand):
                command = chat.commands[command_names[j]](
                    name=command_names[j])
            else:
                command = chat.commands[command_names[j]](
                    cell.value, name=command_names[j])

            # Process CompleteCommand
            if isinstance(command, CompleteCommand):
                idx += 1
                completion_call_id = f"idx {idx}"
                command = CompleteCommand(call_id=join_none(call_id, completion_call_id))

            # Execute command
            chat.execute(command)
            if isinstance(command, CompleteCommand):
                _modify_cell(cell, lock, chat.last_assistant_message())

        except Exception as e:
            exceptions.append(e)
            logger.warning(write_error(e, name, now_delta(start), sep=': '))
            _modify_cell(cell, lock, write_exception(e), error=True)

    # Finish message
    logger.info(f"Completed {name} in {write_now_delta(start)}")

    # Return exceptions or None
    if exceptions: return exceptions
    else: return None


def complete(
        workbook_path: Path | str, new_workbook_path: Path | str = None,
        sheet_names: Sequence[str] | None = None,
        api_keys: dict[type(APIWrapper): str] | None = None,
        api_key_files: dict[type(APIWrapper): Path] | None = None,
        model: str = Chat.MODELS[0],
        timeout: int | None = None, retries: int | None = 0,
        max_chats: int | None = None,
        file_id: str | int | None = None):
    """Complete chats in Excel workbook columns

    Excel sheet format
    - Values starting with COMMENT_SUFFIX are considered comments
    - All columns and all sheets in workbook will be completed by default
      Comment name of sheet or header to skip sheet or column
    - Params are read from the first column with non-commented heading (including None)

    OpenAI API Key must be set prior to function call, see set_openai_api_key()
    """

    start = datetime.now()
    logger = module_logger.getChild(complete.__name__)

    # Load workbook and sheet
    logger.info(
        f"Starting workbook chat completion{rmark(file_id)}: {workbook_path}")
    wb = load_workbook(filename=workbook_path)

    # Additional debug information
    logger.debug(f"File identification (id): {file_id}")
    logger.debug(f"Model{rmark(file_id)}: {model}")

    # Setup sheet_names
    sheet_name_matches = None
    if sheet_names is not None:
        sheet_name_matches = tuple(name.lower() for name in sheet_names)

    # Concurrent completion of all columns
    i_chat = 0
    with ThreadPoolExecutor(max_workers=max_chats) as executor:

        # Workbook lock for modification (for openpyxl)
        lock = Lock()

        # Futures dictionary {future: name}
        future_completes = {}

        # Sheets loop
        i_sheet = -1
        for sheet in wb.worksheets:

            # Test for skipping sheet
            if (
                    (sheet_name_matches and sheet.title.lower() not in sheet_name_matches) or
                    sheet.title.startswith(COMMENT_CHAR)):
                continue
            logger.info(f"Accessing worksheet '{sheet.title}'{rmark(file_id)}")

            # Get params
            command_names = None
            i_command_col = 0
            for i_command_col, col in enumerate(sheet.iter_cols()):
                if str(col[0].value).startswith(COMMENT_CHAR): continue
                i_sheet += 1
                sheet_id = f"chat {i_sheet}"
                command_names = [str(cell.value).strip().lower() for cell in col[1:]]
                logger.info(
                    f"Command sequence{rmark(file_id, join_none(sheet_id, 'x', sep=CHAT_ID_SEP))}: "
                    f"{', '.join(command_names)}")
                break
            if command_names is None:
                logger.info(f"Cannot find command, skipping empty sheet{rmark(file_id)}")
                continue

            # Columns iteration in sheet
            for col in sheet.iter_cols(min_col=i_command_col + 1 + 1):
                if str(col[0].value).startswith(COMMENT_CHAR): continue
                call_id = join_none(
                    file_id, join_none(sheet_id, i_chat, sep=CHAT_ID_SEP))
                name = f"column '{col[0].value}'{rmark(call_id)}"
                logger.info(f"Pooling {name}")
                future = executor.submit(
                    _complete_col, name=name, command_names=command_names, col=col,
                    api_keys=api_keys, api_key_files=api_key_files, model=model,
                    timeout=timeout, retries=retries, lock=lock,
                    call_id=call_id)
                future_completes[future] = name
                i_chat += 1

        # Wait for all future to complete and manage exceptions
        exceptions = []
        for future in as_completed(future_completes, timeout=None):
            if future.exception() is not None:
                logger.critical(
                    write_error(
                        future.exception(), future_completes[future], re_raise=True,
                        sep='\n'))
                raise future.exception()
            elif future.result() is not None:
                exceptions.append(future.result())

    # Save workbook
    if new_workbook_path is None: new_workbook_path = workbook_path
    logger.info(f"Saving workbook{rmark(file_id)}: {new_workbook_path}")
    wb.save(new_workbook_path)

    # Final message
    end = f"workbook with {i_chat} chat(s){rmark(file_id)} in {write_now_delta(start)}"
    if len(exceptions) > 0:
        logger.warning(f"{len(exceptions)} error(s) completing {end}")
    elif i_chat == 0:
        logger.info(f"Finished processing {end}")
    else:
        logger.info(f"Successfully completed {end}")

    # Return
    return exceptions if len(exceptions) > 0 else None
